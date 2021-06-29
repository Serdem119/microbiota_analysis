cluster1= ["NORMAL","NORMAL-NORMAL","OBESE-NORMAL","NORMAL","STOOL","STOOL","NORMAL","NORMAL","OBESE","OBESE"]
cluster2= ["OBESE","NORMAL-OBESE","OBESE-OBESE","OBESE","MECONIUM","MECONIUM","NORMAL-NORMAL","NORMAL-OBESE","OBESE-NORMAL","OBESE-OBESE"]
metadata_path= 'D:\Sample\MİKRO ANALİZ GRUPLANDIRMA.xlsx'
taxa_path= 'D:\Sample'
output_path= "D:\Sample\deneme\\"
def prevalence_calculator(metadata_path,taxa_path,output_path,cluster1, cluster2):  

    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import os
    ##Part1
    "1.1: read metadata"
    metadata=pd.read_excel(metadata_path,sheet_name='Sayfa1')

    "1.2: discard unnecessay columns"
    metadata=metadata.iloc[:,4::]


    """1.3: Separate data by groups w/ sample-id"""
    groupnames=[]

    for i in range(len(metadata.columns)-1):
        groupnames.append('g'+str(i+1))
        globals()[groupnames[i]]=metadata.iloc[:,[0,i+1]]
    
    """1.4: drop N/A values"""

    for i in range(len(groupnames)):
        globals()[groupnames[i]]=globals()[groupnames[i]].dropna()

        
    """1.5: cluster in groups as gn1&&gn2"""

    counter0=0
    groupnew=[]
    exitname=[]
    for grpnm in groupnames:
        counter1=-1
        counter2=-1
        exitname.append(str(globals()[grpnm].columns[1]))
        for i in range(len(globals()[grpnm].iloc[:,1])):
            if globals()[grpnm].iloc[i,1]== cluster1[counter0]:
                counter1 +=1
                counter2 +=1
                globals()[grpnm+"1"]=globals()[grpnm].iloc[0:counter1+1,::]
            
            if globals()[grpnm].iloc[i,1]== cluster2[counter0]:
                counter2 +=1
                globals()[grpnm+"2"]=globals()[grpnm].iloc[counter1+1:counter2+1,::]

        groupnew.append(grpnm+"1")
        groupnew.append(grpnm+"2")
        counter0 +=1


    ##Part2
    """2.1: reach working directory"""

    os.chdir(taxa_path)

    """2.2: import multiple DFs"""
    files=os.listdir(taxa_path)
    files_xls = [f for f in files if f[-13:-5] == 'taxonomy']


    for i in range(len(files_xls)):
        globals()[files_xls[i]]=pd.read_excel(str(os.getcwd())+ "\\"+ files_xls[i])
    
    ##Part3
    for grpnm in groupnew:
        globals()[grpnm+'_df']=pd.DataFrame()
        for i in range(len(globals()[grpnm])):
            globals()[grpnm+'_df']=pd.concat([(globals()[grpnm+'_df']),
                                              (globals()[globals()[grpnm].iloc[i,0]+'taxonomy.xlsx'])])

    """select ~% columns"""
    for grpnm in groupnew:
        globals()[grpnm+'_df']=globals()[grpnm+'_df'].iloc[:,1::2]
    

    ##Part4
    """4.1: get longest column"""
    counter3=0
    wb= Workbook() 
    for grpnm in groupnew:
        globals()[grpnm+'_finaldf']=pd.DataFrame()
        globals()[grpnm+'_free']=pd.DataFrame()
        LP=len((globals()[grpnm+'_df']).pivot_table(index=['Phylum'], aggfunc='size').index)
        LC=len((globals()[grpnm+'_df']).pivot_table(index=['Class'], aggfunc='size').index)
        LO=len((globals()[grpnm+'_df']).pivot_table(index=['Order'], aggfunc='size').index)
        LF=len((globals()[grpnm+'_df']).pivot_table(index=['Family'], aggfunc='size').index)
        LG=len((globals()[grpnm+'_df']).pivot_table(index=['Genus'], aggfunc='size').index)
        LS=len((globals()[grpnm+'_df']).pivot_table(index=['Species'], aggfunc='size').index)
        max_len=max(LP,LC,LO,LF,LG,LS)
    
        """4.2: create DF w/ longest column"""    

    #get taxa && prev values then match w/ max_len    
        globals()[grpnm+'_finaldf']["Phylum"]=list((((globals()[grpnm+'_df'])).pivot_table(index=['Phylum'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LP))
        globals()[grpnm+'_finaldf']["PrevPhylum"]=list((((globals()[grpnm+'_df'])).pivot_table(index=['Phylum'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LP))        
        
        globals()[grpnm+'_finaldf']["Class"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Class'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LC))
        globals()[grpnm+'_finaldf']["PrevClass"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Class'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LC))

        globals()[grpnm+'_finaldf']["Order"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Order'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LO))   
        globals()[grpnm+'_finaldf']["PrevOrder"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Order'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LO))
    
        globals()[grpnm+'_finaldf']["Family"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Family'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LF))
        globals()[grpnm+'_finaldf']["PrevFamily"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Family'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LF))
    
        globals()[grpnm+'_finaldf']["Genus"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Genus'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LG))
        globals()[grpnm+'_finaldf']["PrevGenus"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Genus'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LG))
    
        globals()[grpnm+'_finaldf']["Species"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Species'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LS))
        globals()[grpnm+'_finaldf']["PrevSpecies"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Species'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LS))

    
        """4.3: to Excel"""
    
        if int(grpnm[-1:])%2 != 0 :
            wb= Workbook()    
            ws0=wb.active
            ws0.title= cluster1[counter3]
            ws1 = wb.create_sheet(cluster2[counter3])
        
            for r in dataframe_to_rows(globals()[grpnm+'_finaldf'], index=False, header=True):
                ws0.append(r)
            
        if int(grpnm[-1:])%2 == 0 :       
            for r in dataframe_to_rows(globals()[grpnm+'_finaldf'], index=False, header=True):
                ws1.append(r)
            wb.save(output_path + exitname[counter3] +".xlsx")
            counter3 +=1

prevalence_calculator(metadata_path,taxa_path,output_path,cluster1, cluster2)












