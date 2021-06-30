# -*- coding: utf-8 -*-
"""
Created on Mon Jun 28 15:17:23 2021

@author: sinan
"""

metadata_path= 'D:\Sample\MİKRO ANALİZ GRUPLANDIRMA.xlsx'
taxa_path= 'D:\Sample'
output_path= "D:\Sample\deneme\\"

def prev_calc_auto(metadata_path,taxa_path,output_path):

    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import os
    ##Part1
    "1.1: read metadata"
    metadata=pd.read_excel(metadata_path,sheet_name='Sayfa1')

    "1.2: discard unnecessay columns"
    metadata=metadata.iloc[:,4::]


    """1.3: Separate data by groups w/ sample-id"""
    groupnames=[]

    for i in range(len(metadata.columns)-1):
        groupnames.append('g'+str(i+1))
        globals()[groupnames[i]]=metadata.iloc[:,[0,i+1]]
    
    """1.4: drop N/A values"""
    for i in range(len(groupnames)):
        globals()[groupnames[i]]=globals()[groupnames[i]].dropna()
     
    """1.5: get MainClass names"""    
    exitname=[]
    for grpnm in groupnames:
        exitname.append(str(globals()[grpnm].columns[1]))
    
    """1.6: cluster in groups as g100n,gx00n,..."""
    ct0=0
    groupnew=[]
    for j in range(len(groupnames)):
        ct0=0
        
        for i in range(globals()[groupnames[j]][exitname[j]].nunique()):
        
            globals()[groupnames[j]+'00'+str(i)]=globals()[groupnames[j]][globals()[groupnames[j]][exitname[j]] == globals()[groupnames[j]][exitname[j]].unique()[i]]
            groupnew.append(groupnames[j]+'00'+str(i))
            ct0 +=1
                
    ##Part2
    """2.1: reach working directory"""
    os.chdir(taxa_path)

    """2.2: import multiple DFs"""
    files=os.listdir(taxa_path)
    files_xls = [f for f in files if f[-13:-5] == 'taxonomy']

    for i in range(len(files_xls)):
        globals()[files_xls[i]]=pd.read_excel(str(os.getcwd())+ "\\"+ files_xls[i])
    
    ##Part3
    """2.3: collect taxa under MainClasses as gx00n_df"""
    for grpnm in groupnew:
        globals()[grpnm+'_df']=pd.DataFrame()
        for i in range(len(globals()[grpnm])):
            globals()[grpnm+'_df']=pd.concat([(globals()[grpnm+'_df']),
                                              (globals()[globals()[grpnm].iloc[i,0]+'taxonomy.xlsx'])])
    
    """2.4: select ~% columns"""
    for grpnm in groupnew:
        globals()[grpnm+'_df']=globals()[grpnm+'_df'].iloc[:,1::2]
    
    """3.1: get cluster lengths && fibonacci seq of lenclust"""
    lenclust=[0]
    for i in range(len(groupnames)):
        lenclust.append((globals()[groupnames[i]])[exitname[i]].nunique())

    for i in range(len(lenclust)-1):
        lenclust[i+1]=lenclust[i] + lenclust[i+1]

    """3.2: create list of SubClass names(sheetnames)"""
    sheetnames=[]    
    for i in range(len(groupnames)):
        for j in range(len((globals()[groupnames[i]])[exitname[i]].unique())):
            sheetnames.append((globals()[groupnames[i]])[exitname[i]].unique()[j])
    
    """4.1: get longest column"""
    counter3=0
    counter4 = 0

    for grpnm in groupnew:
        globals()[grpnm+'_finaldf']=pd.DataFrame()
        globals()[grpnm+'_free']=pd.DataFrame()
        LP=len((globals()[grpnm+'_df']).pivot_table(index=['Phylum'], aggfunc='size').index)
        LC=len((globals()[grpnm+'_df']).pivot_table(index=['Class'], aggfunc='size').index)
        LO=len((globals()[grpnm+'_df']).pivot_table(index=['Order'], aggfunc='size').index)
        LF=len((globals()[grpnm+'_df']).pivot_table(index=['Family'], aggfunc='size').index)
        LG=len((globals()[grpnm+'_df']).pivot_table(index=['Genus'], aggfunc='size').index)
        LS=len((globals()[grpnm+'_df']).pivot_table(index=['Species'], aggfunc='size').index)
        max_len=max(LP,LC,LO,LF,LG,LS)
    
        """4.2: create DF w/ longest column"""    

    #get taxa && prev values then match w/ max_len    
        globals()[grpnm+'_finaldf']["Phylum"]=list((((globals()[grpnm+'_df'])).pivot_table(index=['Phylum'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LP))
        globals()[grpnm+'_finaldf']["PrevPhylum"]=list((((globals()[grpnm+'_df'])).pivot_table(index=['Phylum'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LP))        
        
        globals()[grpnm+'_finaldf']["Class"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Class'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LC))
        globals()[grpnm+'_finaldf']["PrevClass"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Class'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LC))

        globals()[grpnm+'_finaldf']["Order"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Order'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LO))   
        globals()[grpnm+'_finaldf']["PrevOrder"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Order'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LO))
    
        globals()[grpnm+'_finaldf']["Family"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Family'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LF))
        globals()[grpnm+'_finaldf']["PrevFamily"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Family'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LF))
    
        globals()[grpnm+'_finaldf']["Genus"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Genus'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LG))
        globals()[grpnm+'_finaldf']["PrevGenus"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Genus'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LG))
    
        globals()[grpnm+'_finaldf']["Species"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Species'], aggfunc='size').sort_values(ascending=False)).index) + (['']*(max_len-LS))
        globals()[grpnm+'_finaldf']["PrevSpecies"]=list(((globals()[grpnm+'_df']).pivot_table(index=['Species'], aggfunc='size').sort_values(ascending=False)).values / len(globals()[grpnm])) + (['']*(max_len-LS))
    
    
        """4.3:create excel workbook && append first sheet // 
            append sheet and save workbook //
            append sheet"""
        if counter3 in lenclust:
            wb=Workbook()
            globals()['ws' + str(counter3)]=wb.active
            globals()['ws' + str(counter3)].title = sheetnames[counter3]
            for r in dataframe_to_rows(globals()[grpnm+'_finaldf'], index=False, header=True):
                globals()['ws' + str(counter3)].append(r)
        
        
        elif counter3 + 1 == lenclust[counter4 + 1]:
            globals()['ws' + str(counter3)] = wb.create_sheet(sheetnames[counter3])
            for r in dataframe_to_rows(globals()[grpnm+'_finaldf'], index=False, header=True):
                globals()['ws' + str(counter3)].append(r)
            
            wb.save(output_path + exitname[counter4] +".xlsx")
            counter4 +=1
    
        else:
            globals()['ws' + str(counter3)] = wb.create_sheet(sheetnames[counter3])
            for r in dataframe_to_rows(globals()[grpnm+'_finaldf'], index=False, header=True):
                globals()['ws' + str(counter3)].append(r)
       
        counter3 +=1
        
        

prev_calc_auto(metadata_path,taxa_path,output_path)
